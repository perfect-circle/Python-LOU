from openpyxl import Workbook, load_workbook
import datetime

wb = load_workbook('courses.xlsx')
students_sheet = wb['students']
time_sheet = wb['time']

def combine():
    courses_sheet = wb.create_sheet('combine')
    courses_sheet.append(['创建时间', '课程名称', '学习人数', '学习时间'])
    for s in students_sheet.values:
        if s[0] != "创建时间":
            for t in time_sheet.values:
                if s[1] == t[1]:
                    courses_sheet.append(list(s) + [t[2]])
    wb.save('courses.xlsx')

def separate():
    combine_sheet = wb['combine']
    time_name = []
    for c in combine_sheet.values:
        if c[0] != '创建时间':
            time_name.append(c[0].strftime('%Y'))

    for name in set(time_name):
        wb_t = Workbook()
        wb_t.remove(wb_t.active)
        time_sheet = wb_t.create_sheet(title=name)
        time_sheet.append(['创建时间', '课程名称', '学习人数', '学习时间'])
        for c in combine_sheet.values:
            if c[0] != '创建时间' and c[0].strftime("%Y") == name:
                time_sheet.append(list(c))
        wb_t.save("{}.xlsx".format(name))

if __name__ == "__main__":
    combine()
    separate()
